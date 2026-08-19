import pandas as pd
from bs4 import BeautifulSoup
import requests
import time
from pathlib import Path
import tkinter as tk
from tkinter import ttk
from openpyxl.styles import Font, PatternFill, Alignment
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ----------------------------
# GUI TKINTER
# ----------------------------
root = tk.Tk()
root.title("Cek Domain")
root.geometry("420x160")
root.resizable(False, False)

label_title = ttk.Label(root, text="Sedang mengecek domain...", font=("Segoe UI", 10, "bold"))
label_title.pack(pady=(15, 5))

progress_var = tk.DoubleVar(value=0)
progress_bar = ttk.Progressbar(root, variable=progress_var, orient="horizontal", length=350, mode="determinate")
progress_bar.pack(pady=5)

label_percent = ttk.Label(root, text="0%")
label_percent.pack(pady=5)

def update_progress(current, total):
    if total == 0:
        percent = 0
    else:
        percent = (current / total) * 100
    progress_var.set(percent)
    label_percent.config(text=f"{int(percent)}%")
    root.update_idletasks()
    root.update()

# ----------------------------
# SELENIUM, BS4
# ----------------------------
opts = Options()
#opts.add_argument("--headless=new")
opts.add_argument("--disable-gpu")
opts.add_argument("--window-size=1920,1080")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
driver.get("http://10.119.105.77:3033")
time.sleep(2)

driver.find_element(By.CSS_SELECTOR, 'input[type="text"]').send_keys("pkl01")
driver.find_element(By.CSS_SELECTOR, 'input[type="password"]').send_keys("pkl01")
driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
time.sleep(2)

list_domain = []

while True:
    html_sumber = driver.page_source
    soup = BeautifulSoup(html_sumber, 'html.parser')
    rows = soup.find_all('tr', id=lambda x: x and x.startswith('row-domain-'))

    for row in rows:
        td_pertama = row.find('td')
        if td_pertama:
            domain = td_pertama.get_text(strip=True)
            if domain not in list_domain:
                list_domain.append(domain)

    next_btn = driver.find_elements(By.ID, "btn-page-next")

    if not next_btn:
        break

    if next_btn[0].get_attribute("disabled") is not None:
        break

    next_btn[0].click()
    time.sleep(2)

driver.quit()

# ----------------------------
# REQUESTS
# ----------------------------
def check_one_domain(domain):
    url = domain if domain.startswith("http") else f"http://{domain}"
    try:
        res = requests.get(url, timeout=5, headers={"User-Agent": "Mozilla/5.0"})
        status = 'Bisa diakses' if res.status_code == 200 else f'HTTP {res.status_code}'
    except Exception:
        status = 'Gagal / Down'
    return {'Domain': domain, 'Status': status}

hasil = [None] * len(list_domain)

with ThreadPoolExecutor(max_workers=40) as executor:
    future_to_index = {
        executor.submit(check_one_domain, domain): index
        for index, domain in enumerate(list_domain)
    }

    completed = 0
    for future in as_completed(future_to_index):
        idx = future_to_index[future]
        hasil[idx] = future.result()
        completed += 1
        update_progress(completed, len(list_domain))

hasil = [item for item in hasil if item is not None]

update_progress(len(list_domain), len(list_domain))
label_title.config(text="Selesai! Menyimpan file Excel...")
root.update()

# ----------------------------
# PANDAS,PATHLIB
# ----------------------------
folder = Path(r"C:\Users\Yusuf\Documents\PKL-FASILKOM\tugas-pkl-main")
folder.mkdir(exist_ok=True)

output_file = folder / "hasil.xlsx"

df = pd.DataFrame(hasil, columns=["Domain", "Status"])

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Domain Check")

    ws = writer.sheets["Domain Check"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

root.destroy()

print(f"Selesai! Hasil disimpan di file: {output_file}")